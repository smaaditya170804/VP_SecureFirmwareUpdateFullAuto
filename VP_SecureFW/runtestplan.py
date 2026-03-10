# -*- coding: utf-8 -*-
r"""
C:\VP_SecureFW\runtestplan.py

Universal testplan runner implementing the exact flows demonstrated by Mithun:

CASE A: SREC present + delivery_method: JFlash
    1) RFP flash SREC
    2) Relay ON
    3) J-Flash update package (jlinkflash.py handles reconnect+readchip+relay-off)
    4) SelfProg flag check with --install

CASE B: delivery_method: SelfProgrammer (NO SREC)
    1) SelfProgrammer download update package
    2) SelfProg flag check with --install

CASE C: delivery_method: InstallOnly (NO SREC)
    1) SelfProg flag check with --install

CASE D: delivery_method: JFlash (NO SREC)
    1) Relay ON
    2) J-Flash update package
    3) SelfProg flag check with --install

CASE E: SREC present + delivery_method: InstallOnly
    1) RFP flash SREC
    2) SelfProg flag check with --install

Usage:
    python runtestplan.py <path_to_yaml_testplan>

Notes:
- Looks for optional product config JSON at:
        productconfigs/<ProductName>.json
    If missing, uses built-in defaults based on your transcripts.
"""
from pathlib import Path
import json
from typing import Tuple
import sys
import subprocess

# --- YAML dependency (PyYAML) ------------------------------------------------
try:
    import yaml  # pip install pyyaml
except ImportError:
    print("[ERROR] PyYAML is required. Please run:  pip install pyyaml")
    sys.exit(1)

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent

def run_cmd(cmd_list, cwd=None, stop_on_error=True, echo=False, suppress_output=True):
    """Run a command (list form). Returns exit code, optionally aborts on error."""
    if echo:
        here = f"(cwd: {cwd})" if cwd else ""
        pretty = " ".join(f'"{c}"' if " " in str(c) else str(c) for c in cmd_list)
        print(f"\n>>> {pretty} {here}".strip())
    
    if suppress_output:
        rc = subprocess.call(cmd_list, cwd=str(cwd) if cwd else None, shell=False, 
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    else:
        rc = subprocess.call(cmd_list, cwd=str(cwd) if cwd else None, shell=False)
    
    if stop_on_error and rc != 0:
        print(f"[ERROR] Command returned {rc}.")
        return rc
    return rc

def load_yaml(yaml_path: Path):
    try:
        with open(yaml_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)
    except Exception as e:
        print(f"[ERROR] Failed to read YAML: {yaml_path}\n{e}")
        sys.exit(2)

def detect_product_name(testplan_name: str) -> str:
    # Heuristic: first token of the 'name' is the product (e.g., "SaverAdvPlus VP Firmware...")
    return (testplan_name or "SaverAdvPlus").split()[0]

def load_product_config(product_name: str) -> dict:
    cfg_path = ROOT / "productconfigs" / f"{product_name}.json"
    if cfg_path.exists():
        with open(cfg_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            return cfg

    # Fallback defaults (from your commands)
    return {
        "product_name": product_name,
        "id_code": "45D6136794F8BA46F7B3435164E0DA29",
        "rfp_device": "RX65x",
        "rfp_tool": "e2l",
        "rfp_interface": "fine",
        "rfp_exe": r"C:\Program Files (x86)\Renesas Electronics\Programming Tools\Renesas Flash Programmer V3.21\rfp-cli.exe",
        "jflash_exe": r"C:\Program Files\SEGGER\JLink\JFlashSPI_CL.exe",
        "jflash_project": r"C:\Software\J-Flash SPI_Projects\V9.12\jlinkflash.jflash",
        "selfprog_exe": r"C:\GBP_Testing\Release_Test\Release_Test\SelfProg_Tool.exe",
        "serial_port": 8,
        "serial_baud": 38400,
        "board_number": 48
    }

# -----------------------------------------------------------------------------
# Flows
# -----------------------------------------------------------------------------
def rfp_flash_srec(cfg: dict, srec_path: str):
    """RFP flash the SREC image (Case A: step 1)."""
    script = ROOT / "rfpflash" / "rfpflash.py"
    cmd = [
        "python", str(script),
        "--rfp", cfg.get("rfp_exe", r"rfp-cli.exe"),
        "--device", cfg.get("rfp_device", "RX65x"),
        "--tool", cfg.get("rfp_tool", "e2l"),
        "--interface", cfg.get("rfp_interface", "fine"),
        "--file", srec_path,
        "--id", cfg.get("id_code", ""),
        "--run"
    ]
    print("  [*] Running RFP flash SREC...")
    rc = run_cmd(cmd, stop_on_error=False)
    if rc != 0:
        print(f"  [ERROR] RFP flash failed with exit code {rc}.")
    return rc

def relay_on():
    """Turn ON J-Link power using your relay control (Case A: step 2)."""
    script = ROOT / "relaycontrol" / "relayon.py"
    cmd = ["python", str(script)]
    print("  [*] Turning relay ON...")
    rc = run_cmd(cmd, stop_on_error=False)
    if rc != 0:
        print(f"  [ERROR] Relay ON failed with exit code {rc}.")
    return rc

def jflash_update(cfg: dict, update_pkg: str):
    """Run jlinkflash.py to program/verify update package (Case A: step 3, Case D step 2)."""
    script = ROOT / "jlinkflash" / "jlinkflash.py"
    cmd = [
        "python", str(script),
        "--jflash", cfg.get("jflash_exe", r"JFlashSPI_CL.exe"),
        "--project", cfg.get("jflash_project", r"jlinkflash.jflash"),
        "--image", update_pkg,
        "--addr", "0x0",
        "--connect",
        "--erasechip",
        "--programverify"
    ]
    print("  [*] Running J-Flash update...")
    rc = run_cmd(cmd, stop_on_error=False)
    if rc != 0:
        print(f"  [ERROR] J-Flash update failed with exit code {rc}.")
    return rc

def selfprog_download(cfg: dict, update_pkg: str):
    """SelfProgrammer download update package (Case B)."""
    script = ROOT / "selfprogrammer" / "selfprogrammerdownload.py"
    cmd = [
        "python", str(script),
        "--exe", cfg.get("selfprog_exe", r"SelfProg_Tool.exe"),
        "--port", str(cfg.get("serial_port", 8)),
        "--baud", str(cfg.get("serial_baud", 38400)),
        "--board", str(cfg.get("board_number", 48)),
        "--bin", update_pkg,
        "--backend"
    ]
    print("  [*] Running SelfProgrammer download...")
    rc = run_cmd(cmd, stop_on_error=False)
    if rc != 0:
        print(f"  [ERROR] SelfProgrammer download failed with exit code {rc}.")
    return rc

def selfprog_flagcheck_install(cfg: dict, expected_flags: dict = None, wait_sec=30, retries=5, interval=2):
    """SelfProg flag check with --install (Case A, B, C, D). Returns (rc, ad_path, ar_path, test_passed)."""
    script = ROOT / "selfprogrammer" / "selfproflagcheck.py"
    logs_dir = ROOT / "selfprogrammer" / "logs"
    # ensure logs directory exists and capture files created by this run
    logs_dir.mkdir(parents=True, exist_ok=True)
    before = set(logs_dir.iterdir())
    cmd = [
        "python", str(script),
        "--exe", cfg.get("selfprog_exe", r"SelfProg_Tool.exe"),
        "--logdir", str(logs_dir),
        "--install",
        "--wait", str(wait_sec),
        "--retries", str(retries),
        "--interval", str(interval),
        "--quiet"
    ]
    print("  [*] Running SelfProg flag check with install...")
    rc = run_cmd(cmd, stop_on_error=False, suppress_output=False)
    # determine which log files were created
    after = set(logs_dir.iterdir())
    new = after - before
    ad_path = None
    ar_path = None
    for p in new:
        name = p.name.lower()
        if "after_download" in name:
            ad_path = p
        elif "after_reset" in name:
            ar_path = p
    if rc != 0:
        print(f"  [ERROR] SelfProg flag check failed with exit code {rc}.")
    
    # Check test pass/fail by comparing flags
    test_passed = True
    if expected_flags and ar_path:
        from selfprogrammer import selfprogflagcheckparser as parser
        ar_text = parser.load_text(ar_path)
        ar_data = parser.parse_pairs(ar_text)
        
        print("\n  [DEBUG] Flag Comparison:")
        for field_key, expected_val in expected_flags.items():
            # Try to find the field in ar_data, handling different case variations
            actual_val = "—"
            for key in ar_data.keys():
                if key.replace(" ", "").lower() == field_key.replace(" ", "").lower():
                    actual_val = ar_data[key]
                    break
            
            actual_normalized = parser.normalize_flag_value(actual_val)
            expected_normalized = parser.normalize_flag_value(str(expected_val))
            matches = actual_normalized == expected_normalized
            status = "✓" if matches else "✗"
            print(f"    {status} {field_key}: expected={expected_normalized}, actual={actual_normalized}")
            
            if not matches:
                test_passed = False
    
    return rc, ad_path, ar_path, test_passed

def execute_test(t: dict, cfg: dict) -> Tuple[bool, bool, str]:
    """
    Execute a single test and return (error_occurred, test_passed, flag_summary).
    """
    tid = t.get("id", "Unknown")
    scenario = t.get("scenario", "")
    method = t.get("delivery_method", "")
    flash_srec = t.get("flash_srec")
    update_pkg = t.get("update_pkg")
    expected_flags = t.get("expected_flags_after_reset", {})

    error_occurred = False
    flag_summary = ""
    test_passed = False

    # CASE A: SREC present + JFlash delivery
    if flash_srec and method == "JFlash":
        if rfp_flash_srec(cfg, flash_srec) != 0:
            error_occurred = True
        if relay_on() != 0:
            error_occurred = True
        if not update_pkg:
            print("  [ERROR] update_pkg is required for JFlash tests.")
            error_occurred = True
        elif jflash_update(cfg, update_pkg) != 0:
            error_occurred = True
        rc, ad, ar, test_passed = selfprog_flagcheck_install(cfg, expected_flags)
        if rc != 0:
            error_occurred = True
        else:
            # capture flag summary text via parser
            import io, contextlib
            from selfprogrammer import selfprogflagcheckparser as parser
            buf = io.StringIO()
            if ad and ar:
                with contextlib.redirect_stdout(buf):
                    parser.summarize(ad, ar, expected_flags if expected_flags else None)
                flag_summary = buf.getvalue()

    # CASE D: JFlash NO SREC
    elif method == "JFlash" and not flash_srec:
        if relay_on() != 0:
            error_occurred = True
        if not update_pkg:
            print("  [ERROR] update_pkg is required for JFlash tests.")
            error_occurred = True
        elif jflash_update(cfg, update_pkg) != 0:
            error_occurred = True
        rc, ad, ar, test_passed = selfprog_flagcheck_install(cfg, expected_flags)
        if rc != 0:
            error_occurred = True
        else:
            import io, contextlib
            from selfprogrammer import selfprogflagcheckparser as parser
            buf = io.StringIO()
            if ad and ar:
                with contextlib.redirect_stdout(buf):
                    parser.summarize(ad, ar, expected_flags if expected_flags else None)
                flag_summary = buf.getvalue()

    # CASE B: SelfProgrammer (NO SREC)
    elif method == "SelfProgrammer":
        if not update_pkg:
            print("  [ERROR] update_pkg is required for SelfProgrammer tests.")
            error_occurred = True
        else:
            if selfprog_download(cfg, update_pkg) != 0:
                error_occurred = True
            rc, ad, ar, test_passed = selfprog_flagcheck_install(cfg, expected_flags)
            if rc != 0:
                error_occurred = True
            else:
                import io, contextlib
                from selfprogrammer import selfprogflagcheckparser as parser
                buf = io.StringIO()
                if ad and ar:
                    with contextlib.redirect_stdout(buf):
                        parser.summarize(ad, ar, expected_flags if expected_flags else None)
                    flag_summary = buf.getvalue()

    # CASE E: SREC present + InstallOnly
    elif flash_srec and method == "InstallOnly":
        if rfp_flash_srec(cfg, flash_srec) != 0:
            error_occurred = True
        rc, ad, ar, test_passed = selfprog_flagcheck_install(cfg, expected_flags)
        if rc != 0:
            error_occurred = True
        else:
            import io, contextlib
            from selfprogrammer import selfprogflagcheckparser as parser
            buf = io.StringIO()
            if ad and ar:
                with contextlib.redirect_stdout(buf):
                    parser.summarize(ad, ar, expected_flags if expected_flags else None)
                flag_summary = buf.getvalue()

    # CASE C: InstallOnly (NO SREC)
    elif method == "InstallOnly":
        rc, ad, ar, test_passed = selfprog_flagcheck_install(cfg, expected_flags)
        if rc != 0:
            error_occurred = True
        else:
            import io, contextlib
            from selfprogrammer import selfprogflagcheckparser as parser
            buf = io.StringIO()
            if ad and ar:
                with contextlib.redirect_stdout(buf):
                    parser.summarize(ad, ar, expected_flags if expected_flags else None)
                flag_summary = buf.getvalue()

    else:
        # Guard rails: if a test gives SREC but not JFlash or InstallOnly, or unknown method
        if flash_srec and method not in ["JFlash", "InstallOnly"]:
            print(f"  [ERROR] SREC provided but delivery_method is '{method}'. Expected 'JFlash' or 'InstallOnly'.")
        else:
            print(f"  [ERROR] Unsupported or missing delivery_method: '{method}'")
        error_occurred = True

    return error_occurred, test_passed, flag_summary

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main():
    import argparse
    
    ap = argparse.ArgumentParser(
        description="Run test plans from YAML with optional filtering",
        usage="python runtestplan.py <path_to_yaml_testplan> [--only-test TEST_ID] [--start-from TEST_ID]"
    )
    ap.add_argument("testplan", help="Path to the YAML testplan file")
    ap.add_argument("--only-test", dest="only_test", default=None, help="Run only a specific test ID")
    ap.add_argument("--start-from", dest="start_from", default=None, help="Start from a specific test ID and run remaining tests")
    
    args = ap.parse_args()
    
    testplan_path = Path(args.testplan).resolve()
    if not testplan_path.exists():
        print(f"[ERROR] Testplan not found: {testplan_path}")
        sys.exit(2)

    tp = load_yaml(testplan_path)
    product_name = detect_product_name(tp.get("name", "SaverAdvPlus"))
    cfg = load_product_config(product_name)

    tests = tp.get("tests", [])
    if not tests:
        print("[WARN] No tests found in YAML.")
        sys.exit(0)
    
    # Filter tests based on options
    if args.only_test:
        tests = [t for t in tests if t.get("id") == args.only_test]
        if not tests:
            print(f"[ERROR] Test ID '{args.only_test}' not found in testplan.")
            sys.exit(2)
        print(f"[INFO] Running only test: {args.only_test}\n")
    elif args.start_from:
        start_idx = None
        for idx, t in enumerate(tp.get("tests", [])):
            if t.get("id") == args.start_from:
                start_idx = idx
                break
        if start_idx is None:
            print(f"[ERROR] Test ID '{args.start_from}' not found in testplan.")
            sys.exit(2)
        tests = tp.get("tests", [])[start_idx:]
        print(f"[INFO] Starting from test: {args.start_from}\n")

    # collect results for report
    results = []
    overall_error = False
    last_srec_test_idx = None
    test_idx = 0
    last_failed_no_srec_idx = None
    last_failed_no_srec_retry_count = 0

    while test_idx < len(tests):
        t = tests[test_idx]
        tid = t.get("id", "Unknown")
        scenario = t.get("scenario", "")
        method = t.get("delivery_method", "")
        flash_srec = t.get("flash_srec")
        update_pkg = t.get("update_pkg")
        expected_flags = t.get("expected_flags_after_reset", {})

        # Track the last test that had flash_srec
        if flash_srec:
            last_srec_test_idx = test_idx

        print(f"\n>>> TEST: {tid}")

        # Execute test with retry logic
        retry_count = 0
        error_occurred = False
        test_passed = False
        flag_summary = ""

        while retry_count < 3:
            error_occurred, test_passed, flag_summary = execute_test(t, cfg)
            
            if not error_occurred:
                break
            
            retry_count += 1
            if retry_count < 3:
                print(f"  [RETRY] Test failed, retrying ({retry_count}/3)...")

        # Print test result with pass/fail status
        if error_occurred:
            print(f">>> TEST: {tid} - COMPLETED WITH ERRORS\n")
            
            # If this test has no flash_srec and it errored, restart from last srec test with 3 cycle retries
            if not flash_srec and last_srec_test_idx is not None:
                # Check if this is the same failed test as before
                if last_failed_no_srec_idx == test_idx:
                    last_failed_no_srec_retry_count += 1
                    if last_failed_no_srec_retry_count < 3:
                        print(f"  [INFO] Test {tid} failed again. Restarting from last SREC test (cycle {last_failed_no_srec_retry_count + 1}/3)...")
                        test_idx = last_srec_test_idx
                        overall_error = True
                        continue
                    else:
                        print(f"  [INFO] Test {tid} failed 3 times. Moving to next test.")
                        last_failed_no_srec_idx = None
                        last_failed_no_srec_retry_count = 0
                else:
                    # First failure for this test without srec
                    last_failed_no_srec_idx = test_idx
                    last_failed_no_srec_retry_count = 1
                    print(f"  [INFO] No SREC in this test. Restarting from last SREC test (cycle 1/3)...")
                    test_idx = last_srec_test_idx
                    overall_error = True
                    continue
        else:
            status_symbol = "✓ PASSED" if test_passed else "✗ FAILED"
            print(f">>> TEST: {tid} - {status_symbol}\n")
            # Reset failed no-srec tracking when a test passes
            if not flash_srec:
                last_failed_no_srec_idx = None
                last_failed_no_srec_retry_count = 0

        # record result
        results.append({
            "id": tid,
            "scenario": scenario,
            "delivery": method,
            "flash_srec": flash_srec or "",
            "update_pkg": update_pkg or "",
            "error": error_occurred,
            "test_passed": test_passed,
            "flag_summary": flag_summary,
            "expected_flags": expected_flags,
        })
        if error_occurred:
            overall_error = True

        test_idx += 1

    # if every test succeeded, create report and clean logs
    if not overall_error:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("[WARN] openpyxl not installed; skipping report generation.")
        else:
            reports_dir = ROOT / "reports"
            reports_dir.mkdir(parents=True, exist_ok=True)
            ts = __import__("datetime").datetime.now().strftime("%Y%m%d-%H%M%S")
            report_name = f"{testplan_path.stem}_{ts}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            
            # Setup styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            flag_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            flag_header_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Main headers
            headers = ["Test ID", "Scenario", "Delivery", "Flash SREC", "Update Package", "Error", "Status"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            
            # Set initial column widths for main data
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 12
            
            # Rows with main test data
            row_num = 2
            for r in results:
                ws[f'A{row_num}'].value = r["id"]
                ws[f'B{row_num}'].value = r["scenario"]
                ws[f'C{row_num}'].value = r["delivery"]
                ws[f'D{row_num}'].value = r["flash_srec"]
                ws[f'E{row_num}'].value = r["update_pkg"]
                ws[f'F{row_num}'].value = "Yes" if r["error"] else "No"
                ws[f'G{row_num}'].value = "✓ PASSED" if r.get("test_passed", False) else "✗ FAILED"
                
                # Color code the status column
                status_cell = ws[f'G{row_num}']
                if r.get("test_passed", False):
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(bold=True, color="006100")
                else:
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(bold=True, color="9C0006")
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    cell = ws[f'{col}{row_num}']
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = thin_border
                
                row_num += 1
            
            # Create flag summary details sheet
            flag_sheet = wb.create_sheet("Flag Summary")
            flag_row = 1
            
            for r in results:
                # Test ID header for this test
                test_id_cell = flag_sheet[f'A{flag_row}']
                test_id_cell.value = f"Test ID: {r['id']}"
                test_id_cell.font = Font(bold=True, size=12)
                test_id_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                flag_sheet.merge_cells(f'A{flag_row}:E{flag_row}')
                test_id_cell.border = thin_border
                flag_row += 1
                
                # Parse flag summary data if available
                if r["flag_summary"]:
                    lines = r["flag_summary"].strip().split('\n')
                    # Extract field-value pairs from the flag summary text
                    in_table = False
                    for line in lines:
                        line_stripped = line.strip()
                        if not line_stripped or line_stripped.startswith('+'):
                            continue
                        if line_stripped.startswith('|'):
                            # Extract values from table row
                            parts = [p.strip() for p in line_stripped.split('|')]
                            parts = [p for p in parts if p]  # Remove empty strings
                            
                            if len(parts) == 3:
                                # This is a data row: | Field | After Download | After Reset |
                                field_name = parts[0]
                                val_download = parts[1]
                                val_reset = parts[2]
                                
                                # Get expected value if available
                                expected_val = "—"
                                if r.get("expected_flags"):
                                    expected_val = str(r["expected_flags"].get(field_name, "—"))
                                
                                flag_sheet[f'A{flag_row}'].value = field_name
                                flag_sheet[f'B{flag_row}'].value = val_download
                                flag_sheet[f'C{flag_row}'].value = val_reset
                                flag_sheet[f'D{flag_row}'].value = expected_val
                                
                                # Style the cells
                                for col in ['A', 'B', 'C', 'D']:
                                    cell = flag_sheet[f'{col}{flag_row}']
                                    cell.border = thin_border
                                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                                
                                flag_row += 1
                else:
                    flag_sheet[f'A{flag_row}'].value = "(No flag data available)"
                    flag_sheet[f'A{flag_row}'].font = Font(italic=True, color="808080")
                    flag_row += 1
                
                # Add blank row for separation
                flag_row += 1
            
            # Set column widths for flag summary sheet
            flag_sheet.column_dimensions['A'].width = 25
            flag_sheet.column_dimensions['B'].width = 20
            flag_sheet.column_dimensions['C'].width = 20
            flag_sheet.column_dimensions['D'].width = 20
            
            wb.save(str(reports_dir / report_name))
            print(f"[INFO] Report generated: {reports_dir / report_name}")
            # cleanup log files now that report exists
            for logdir in ROOT.rglob('logs'):
                if logdir.is_dir():
                    for f in logdir.iterdir():
                        try:
                            f.unlink()
                        except Exception:
                            pass
            print("[INFO] Log files deleted.")
    print("\n" + "="*60)
    print("ALL TESTS COMPLETED")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()